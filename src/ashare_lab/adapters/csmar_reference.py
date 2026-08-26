"""Read the licensed CSMAR balance-sheet and index ZIPs without modifying them.

This adapter deliberately keeps balance-sheet rows separate from the historical
factor pipeline.  The supplied ``FS_Combas`` export has accounting periods and
correction dates, but it does not contain the ordinary publication date needed
to know when a historical investor could first have observed each statement.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import tempfile
import zipfile
from collections.abc import Iterator
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import BinaryIO, Literal

import pandas as pd
from openpyxl import load_workbook

from ashare_lab.adapters.csmar_local import infer_a_share_exchange
from ashare_lab.domain.errors import DataQualityError, DataUnavailableError

BALANCE_SHEET_FIELDS = (
    "Stkcd",
    "ShortName",
    "Accper",
    "Typrep",
    "IfCorrect",
    "DeclareDate",
    "A001101000",  # cash and cash equivalents reported as monetary funds
    "A001111000",  # accounts receivable, net
    "A001123000",  # inventory, net
    "A001100000",  # current assets
    "A001000000",  # total assets
    "A002100000",  # current liabilities
    "A002000000",  # total liabilities
    "A003100000",  # equity attributable to the parent
    "A003000000",  # total owners' equity
)

INDEX_DAILY_FIELDS = (
    "Indexcd",
    "Idxtrd01",
    "Idxtrd02",
    "Idxtrd03",
    "Idxtrd04",
    "Idxtrd05",
    "Idxtrd06",
    "Idxtrd07",
    "Idxtrd08",
)

BALANCE_OUTPUT_COLUMNS = (
    "symbol",
    "name",
    "report_period",
    "report_type",
    "is_correction",
    "correction_disclosure_dates",
    "cash_cny",
    "accounts_receivable_cny",
    "inventory_cny",
    "current_assets_cny",
    "total_assets_cny",
    "current_liabilities_cny",
    "total_liabilities_cny",
    "parent_equity_cny",
    "total_equity_cny",
    "ordinary_announcement_date",
    "data_role",
    "historical_backtest_eligible",
    "common_cutoff_date",
    "source",
    "retrieved_at",
)

INDEX_OUTPUT_COLUMNS = (
    "index_code",
    "trade_date",
    "open",
    "high",
    "low",
    "close",
    "component_volume",
    "component_amount_cny",
    "index_return",
    "knowledge_date",
    "data_role",
    "historical_backtest_eligible",
    "common_cutoff_date",
    "source",
    "retrieved_at",
)


@dataclass(frozen=True, slots=True)
class CSMARReferenceEntry:
    kind: Literal["balance_sheet", "index_daily"]
    archive_path: Path
    member_name: str
    archive_sha256: str
    crc32: int
    file_size: int
    compressed_size: int

    @property
    def entry_id(self) -> str:
        payload = f"{self.kind}\0{self.archive_path.name}\0{self.member_name}"
        digest = hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]
        return f"{self.kind}-{digest}"

    @property
    def display_name(self) -> str:
        return f"{self.archive_path.name}:{self.member_name}"


@dataclass(frozen=True, slots=True)
class CSMARReferenceLayout:
    root: Path
    balance_sheet_entry: CSMARReferenceEntry
    index_entries: tuple[CSMARReferenceEntry, ...]
    fingerprint: str

    @property
    def entries(self) -> tuple[CSMARReferenceEntry, ...]:
        return (self.balance_sheet_entry, *self.index_entries)


class CSMARReferenceReader:
    """Streaming reader for the extra CSMAR ZIPs supplied by the user."""

    def __init__(self, source_root: str | Path) -> None:
        self.layout = discover_csmar_reference_layout(source_root)

    def iter_raw_batches(
        self,
        entry: CSMARReferenceEntry,
        *,
        batch_size: int,
    ) -> Iterator[pd.DataFrame]:
        if batch_size <= 0:
            raise ValueError("batch_size必须大于0")
        fields = BALANCE_SHEET_FIELDS if entry.kind == "balance_sheet" else INDEX_DAILY_FIELDS
        rows: list[tuple[object, ...]] = []
        with _open_nested_workbook(entry) as workbook:
            for row in _iter_workbook_rows(workbook, fields):
                rows.append(row)
                if len(rows) >= batch_size:
                    yield pd.DataFrame(rows, columns=list(fields))
                    rows = []
        if rows:
            yield pd.DataFrame(rows, columns=list(fields))


class CSMARReferenceData:
    """Query a completed reference-data import without exposing raw CSMAR ZIPs."""

    def __init__(self, dataset_root: str | Path) -> None:
        self.root = Path(dataset_root).expanduser().resolve()
        self.database_path = self.root / "csmar_reference.duckdb"

    def read_balance_sheet_snapshot(
        self,
        *,
        symbols: tuple[str, ...] | None = None,
    ) -> pd.DataFrame:
        connection = self._connect()
        try:
            if symbols:
                normalized = tuple(_normalize_security_code(symbol) for symbol in symbols)
                placeholders = ", ".join("?" for _ in normalized)
                result = connection.execute(
                    f"SELECT * FROM balance_sheet_snapshot WHERE symbol IN ({placeholders}) "
                    "ORDER BY symbol, report_period",
                    list(normalized),
                ).fetchdf()
            else:
                result = connection.execute(
                    "SELECT * FROM balance_sheet_snapshot ORDER BY symbol, report_period"
                ).fetchdf()
        finally:
            connection.close()
        if result.empty:
            raise DataUnavailableError("CSMAR资产负债表快照为空")
        if bool(result["historical_backtest_eligible"].any()):
            raise DataQualityError("资产负债表快照被错误标记为可做历史回测")
        return result

    def read_index_daily(
        self,
        index_code: str,
        start: date,
        end: date,
    ) -> pd.DataFrame:
        if start > end:
            raise ValueError("start不能晚于end")
        normalized = _normalize_index_code(index_code)
        connection = self._connect()
        try:
            result = connection.execute(
                "SELECT * FROM index_daily WHERE index_code = ? "
                "AND trade_date BETWEEN ? AND ? ORDER BY trade_date",
                [normalized, start, end],
            ).fetchdf()
        finally:
            connection.close()
        if result.empty:
            raise DataUnavailableError(
                f"CSMAR指数库在{start.isoformat()}至{end.isoformat()}没有{normalized}"
            )
        if not bool(result["historical_backtest_eligible"].all()):
            raise DataQualityError("指数日线被错误标记为不可做历史回测")
        return result

    def _connect(self):
        if not self.database_path.is_file():
            raise DataUnavailableError(f"CSMAR参考数据库不存在：{self.database_path}")
        try:
            import duckdb
        except ImportError as exc:
            raise DataUnavailableError("缺少duckdb依赖，无法查询CSMAR参考数据库") from exc
        return duckdb.connect(str(self.database_path), read_only=True)


def discover_csmar_reference_layout(source_root: str | Path) -> CSMARReferenceLayout:
    root = Path(source_root).expanduser().resolve()
    if not root.is_dir():
        raise DataUnavailableError(f"CSMAR目录不存在：{root}")

    finance_archive = root / "财务数据.zip"
    if not finance_archive.is_file():
        raise DataUnavailableError(f"CSMAR目录缺少财务数据.zip：{root}")
    index_archives = sorted(
        (path.resolve() for path in (root / "大盘数据").glob("*.zip")),
        key=_natural_path_key,
    )
    if not index_archives:
        raise DataUnavailableError("CSMAR目录缺少大盘数据/*.zip")

    archive_hashes = {
        archive: _sha256_file(archive) for archive in (finance_archive.resolve(), *index_archives)
    }
    balance_entries = _discover_entries(
        finance_archive.resolve(),
        kind="balance_sheet",
        pattern=r"FS_Combas\.xlsx",
        archive_sha256=archive_hashes[finance_archive.resolve()],
    )
    if len(balance_entries) != 1:
        raise DataQualityError(
            f"财务数据.zip必须且只能包含一个FS_Combas.xlsx，实际找到{len(balance_entries)}个"
        )

    index_entries: list[CSMARReferenceEntry] = []
    for archive in index_archives:
        discovered = _discover_entries(
            archive,
            kind="index_daily",
            pattern=r"IDX_Idxtrd\d*\.xlsx",
            archive_sha256=archive_hashes[archive],
        )
        if not discovered:
            raise DataQualityError(f"{archive.name}没有IDX_Idxtrd*.xlsx")
        index_entries.extend(discovered)

    fingerprint_payload = {
        "schema": 1,
        "entries": [
            {
                "kind": entry.kind,
                "archive": str(entry.archive_path.relative_to(root)),
                "archive_sha256": entry.archive_sha256,
                "member": entry.member_name,
                "crc32": entry.crc32,
                "file_size": entry.file_size,
                "compressed_size": entry.compressed_size,
            }
            for entry in (balance_entries[0], *index_entries)
        ],
    }
    fingerprint = hashlib.sha256(
        json.dumps(fingerprint_payload, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return CSMARReferenceLayout(
        root=root,
        balance_sheet_entry=balance_entries[0],
        index_entries=tuple(index_entries),
        fingerprint=fingerprint,
    )


def normalize_balance_sheet_snapshot(
    raw: pd.DataFrame,
    *,
    common_cutoff_date: date,
    retrieved_at: date,
) -> pd.DataFrame:
    _require_columns(raw, BALANCE_SHEET_FIELDS, dataset="CSMAR资产负债表")
    frame = raw.loc[:, list(BALANCE_SHEET_FIELDS)].copy()
    frame["symbol"] = frame["Stkcd"].map(_normalize_security_code)
    valid_code = frame["symbol"].str.fullmatch(r"\d{6}", na=False)
    if not bool(valid_code.all()):
        examples = frame.loc[~valid_code, "Stkcd"].head(5).tolist()
        raise DataQualityError(f"CSMAR资产负债表存在无效证券代码：{examples}")
    frame = frame.loc[frame["symbol"].map(infer_a_share_exchange).notna()].copy()
    frame["report_period"] = pd.to_datetime(frame["Accper"], errors="coerce").dt.normalize()
    if bool(frame["report_period"].isna().any()):
        examples = frame.loc[frame["report_period"].isna(), ["symbol", "Accper"]].head(5)
        raise DataQualityError(
            f"CSMAR资产负债表存在无效统计截止日期：{examples.to_dict('records')}"
        )

    report_type = frame["Typrep"].fillna("").astype(str).str.strip().str.upper()
    frame = frame.loc[report_type.eq("A")].copy()
    is_opening_balance = frame["report_period"].dt.strftime("%m-%d").eq("01-01")
    frame = frame.loc[~is_opening_balance].copy()
    frame = frame.loc[frame["report_period"] <= pd.Timestamp(common_cutoff_date)].copy()
    if frame.empty:
        return pd.DataFrame(columns=list(BALANCE_OUTPUT_COLUMNS))

    numeric_map = {
        "A001101000": "cash_cny",
        "A001111000": "accounts_receivable_cny",
        "A001123000": "inventory_cny",
        "A001100000": "current_assets_cny",
        "A001000000": "total_assets_cny",
        "A002100000": "current_liabilities_cny",
        "A002000000": "total_liabilities_cny",
        "A003100000": "parent_equity_cny",
        "A003000000": "total_equity_cny",
    }
    for raw_name, canonical_name in numeric_map.items():
        frame[canonical_name] = pd.to_numeric(frame[raw_name], errors="coerce")
    if bool((frame["total_assets_cny"].dropna() < 0).any()):
        raise DataQualityError("CSMAR资产负债表存在负的资产总计")

    correction_text = frame["IfCorrect"].fillna("").astype(str).str.strip().str.lower()
    correction_numeric = pd.to_numeric(frame["IfCorrect"], errors="coerce")
    output = pd.DataFrame(
        {
            "symbol": frame["symbol"],
            "name": frame["ShortName"].fillna("").astype(str).str.strip(),
            "report_period": frame["report_period"],
            "report_type": "A",
            "is_correction": correction_numeric.eq(1) | correction_text.isin(("true", "yes", "是")),
            "correction_disclosure_dates": frame["DeclareDate"].fillna("").astype(str).str.strip(),
            "cash_cny": frame["cash_cny"],
            "accounts_receivable_cny": frame["accounts_receivable_cny"],
            "inventory_cny": frame["inventory_cny"],
            "current_assets_cny": frame["current_assets_cny"],
            "total_assets_cny": frame["total_assets_cny"],
            "current_liabilities_cny": frame["current_liabilities_cny"],
            "total_liabilities_cny": frame["total_liabilities_cny"],
            "parent_equity_cny": frame["parent_equity_cny"],
            "total_equity_cny": frame["total_equity_cny"],
            # Deliberately unknown: DeclareDate is only the correction date.
            "ordinary_announcement_date": pd.NaT,
            "data_role": "current_snapshot",
            "historical_backtest_eligible": False,
            "common_cutoff_date": pd.Timestamp(common_cutoff_date),
            "source": "csmar:FS_Combas",
            "retrieved_at": retrieved_at.isoformat(),
        }
    )
    return (
        output.loc[:, list(BALANCE_OUTPUT_COLUMNS)]
        .sort_values(["symbol", "report_period"], kind="stable")
        .reset_index(drop=True)
    )


def normalize_index_daily(
    raw: pd.DataFrame,
    *,
    common_cutoff_date: date,
    retrieved_at: date,
) -> pd.DataFrame:
    _require_columns(raw, INDEX_DAILY_FIELDS, dataset="CSMAR指数日线")
    frame = raw.loc[:, list(INDEX_DAILY_FIELDS)].copy()
    frame["index_code"] = frame["Indexcd"].map(_normalize_index_code)
    valid_code = frame["index_code"].str.fullmatch(r"[A-Za-z0-9._-]+", na=False)
    if not bool(valid_code.all()):
        examples = frame.loc[~valid_code, "Indexcd"].head(5).tolist()
        raise DataQualityError(f"CSMAR指数日线存在无效指数代码：{examples}")
    frame["trade_date"] = pd.to_datetime(frame["Idxtrd01"], errors="coerce").dt.normalize()
    if bool(frame["trade_date"].isna().any()):
        examples = frame.loc[frame["trade_date"].isna(), ["index_code", "Idxtrd01"]].head(5)
        raise DataQualityError(f"CSMAR指数日线存在无效日期：{examples.to_dict('records')}")
    frame = frame.loc[frame["trade_date"] <= pd.Timestamp(common_cutoff_date)].copy()
    if frame.empty:
        return pd.DataFrame(columns=list(INDEX_OUTPUT_COLUMNS))

    numeric_map = {
        "Idxtrd02": "open",
        "Idxtrd03": "high",
        "Idxtrd04": "low",
        "Idxtrd05": "close",
        "Idxtrd06": "component_volume",
        "Idxtrd07": "component_amount_cny",
        "Idxtrd08": "index_return",
    }
    for raw_name, canonical_name in numeric_map.items():
        frame[canonical_name] = pd.to_numeric(frame[raw_name], errors="coerce")
    invalid_price = frame[["open", "high", "low", "close"]].isna().any(axis=1) | (
        frame[["open", "high", "low", "close"]] < 0
    ).any(axis=1)
    invalid_price_count = int(invalid_price.sum())
    frame = frame.loc[~invalid_price].copy()
    if frame.empty:
        empty = pd.DataFrame(columns=list(INDEX_OUTPUT_COLUMNS))
        empty.attrs["invalid_price_rows_dropped"] = invalid_price_count
        empty.attrs["invalid_geometry_rows_dropped"] = 0
        return empty
    invalid_geometry = frame["high"] < frame[["open", "low", "close"]].max(axis=1)
    invalid_geometry |= frame["low"] > frame[["open", "high", "close"]].min(axis=1)
    invalid_geometry_count = int(invalid_geometry.sum())
    # The supplied CSMAR archive contains a material number of old index rows
    # whose OHLC geometry is internally impossible (not mere float epsilon).
    # Quarantine them with an explicit audit count; never repair or silently
    # reinterpret the vendor values.
    frame = frame.loc[~invalid_geometry].copy()
    if frame.empty:
        empty = pd.DataFrame(columns=list(INDEX_OUTPUT_COLUMNS))
        empty.attrs["invalid_price_rows_dropped"] = invalid_price_count
        empty.attrs["invalid_geometry_rows_dropped"] = invalid_geometry_count
        empty.attrs["invalid_activity_rows_dropped"] = 0
        return empty
    invalid_activity = (frame[["component_volume", "component_amount_cny"]] < 0).any(axis=1)
    invalid_activity_count = int(invalid_activity.sum())
    frame = frame.loc[~invalid_activity].copy()
    if frame.empty:
        empty = pd.DataFrame(columns=list(INDEX_OUTPUT_COLUMNS))
        empty.attrs["invalid_price_rows_dropped"] = invalid_price_count
        empty.attrs["invalid_geometry_rows_dropped"] = invalid_geometry_count
        empty.attrs["invalid_activity_rows_dropped"] = invalid_activity_count
        return empty

    output = pd.DataFrame(
        {
            "index_code": frame["index_code"],
            "trade_date": frame["trade_date"],
            "open": frame["open"],
            "high": frame["high"],
            "low": frame["low"],
            "close": frame["close"],
            "component_volume": frame["component_volume"],
            "component_amount_cny": frame["component_amount_cny"],
            "index_return": frame["index_return"],
            "knowledge_date": frame["trade_date"],
            "data_role": "historical_point_in_time",
            "historical_backtest_eligible": True,
            "common_cutoff_date": pd.Timestamp(common_cutoff_date),
            "source": "csmar:IDX_Idxtrd",
            "retrieved_at": retrieved_at.isoformat(),
        }
    )
    result = (
        output.loc[:, list(INDEX_OUTPUT_COLUMNS)]
        .sort_values(["index_code", "trade_date"], kind="stable")
        .reset_index(drop=True)
    )
    result.attrs["invalid_price_rows_dropped"] = invalid_price_count
    result.attrs["invalid_geometry_rows_dropped"] = invalid_geometry_count
    result.attrs["invalid_activity_rows_dropped"] = invalid_activity_count
    return result


def _discover_entries(
    archive_path: Path,
    *,
    kind: Literal["balance_sheet", "index_daily"],
    pattern: str,
    archive_sha256: str,
) -> list[CSMARReferenceEntry]:
    try:
        with zipfile.ZipFile(archive_path) as bundle:
            infos = [
                info
                for info in bundle.infolist()
                if re.fullmatch(pattern, Path(info.filename).name)
            ]
    except zipfile.BadZipFile as exc:
        raise DataQualityError(f"CSMAR压缩包损坏：{archive_path}") from exc
    infos.sort(key=lambda info: _natural_text_key(Path(info.filename).name))
    return [
        CSMARReferenceEntry(
            kind=kind,
            archive_path=archive_path,
            member_name=info.filename,
            archive_sha256=archive_sha256,
            crc32=info.CRC,
            file_size=info.file_size,
            compressed_size=info.compress_size,
        )
        for info in infos
    ]


@contextmanager
def _open_nested_workbook(entry: CSMARReferenceEntry):
    with zipfile.ZipFile(entry.archive_path) as bundle:
        try:
            source: BinaryIO = bundle.open(entry.member_name)
        except KeyError as exc:
            raise DataQualityError(f"压缩包成员消失：{entry.display_name}") from exc
        with source, tempfile.SpooledTemporaryFile(max_size=64 * 1024 * 1024) as staged:
            shutil.copyfileobj(source, staged, length=1024 * 1024)
            staged.seek(0)
            workbook = load_workbook(staged, read_only=True, data_only=True)
            try:
                yield workbook
            finally:
                workbook.close()


def _iter_workbook_rows(workbook, required_fields: tuple[str, ...]):
    worksheet = workbook.active
    worksheet.reset_dimensions()
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header = tuple(str(value).strip() if value is not None else "" for value in next(iterator))
    except StopIteration as exc:
        raise DataUnavailableError("CSMAR工作簿为空") from exc
    missing = set(required_fields) - set(header)
    if missing:
        raise DataQualityError("CSMAR工作簿缺少字段：" + ", ".join(sorted(missing)))
    positions = tuple(header.index(field) for field in required_fields)
    try:
        next(iterator)
        next(iterator)
    except StopIteration as exc:
        raise DataUnavailableError("CSMAR工作簿只有表头，没有数据") from exc
    for row in iterator:
        if not row or all(value is None for value in row):
            continue
        yield tuple(row[position] if position < len(row) else None for position in positions)


def _require_columns(raw: pd.DataFrame, required: tuple[str, ...], *, dataset: str) -> None:
    missing = set(required) - set(raw.columns)
    if missing:
        raise DataQualityError(f"{dataset}缺少字段：" + ", ".join(sorted(missing)))


def _normalize_security_code(value: object) -> str:
    text = _plain_code(value)
    for suffix in (".SH", ".SZ", ".BJ"):
        if text.upper().endswith(suffix):
            text = text[: -len(suffix)]
            break
    return text.zfill(6) if text.isdigit() and len(text) <= 6 else text


def _normalize_index_code(value: object) -> str:
    text = _plain_code(value)
    return text.zfill(6) if text.isdigit() and len(text) <= 6 else text.upper()


def _plain_code(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") and text[:-2].isdigit() else text


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while block := handle.read(1024 * 1024):
            digest.update(block)
    return digest.hexdigest()


def _natural_path_key(path: Path) -> tuple[object, ...]:
    return _natural_text_key(path.name)


def _natural_text_key(value: str) -> tuple[object, ...]:
    return tuple(int(part) if part.isdigit() else part for part in re.split(r"(\d+)", value))
