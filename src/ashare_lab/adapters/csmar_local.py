"""Read a user-supplied CSMAR Excel export without touching the originals.

CSMAR writes an incorrect ``A1:A...`` worksheet dimension in some exports.
``openpyxl`` trusts that metadata in read-only mode, so every worksheet must
call ``reset_dimensions`` before iteration or only the first column is read.
"""

from __future__ import annotations

import hashlib
import json
import re
import shutil
import tempfile
import zipfile
from collections.abc import Iterator
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import UTC, date, datetime
from pathlib import Path
from typing import BinaryIO

import pandas as pd
from openpyxl import load_workbook

from ashare_lab.domain.errors import DataQualityError, DataUnavailableError
from ashare_lab.ports.market_data import CANONICAL_DAILY_COLUMNS, normalize_symbol

CSMAR_MASTER_FIELDS = (
    "Stkcd",
    "Stknme",
    "Listdt",
    "Indnme",
    "Nindnme",
    "Nnindnme",
    "IndnmeZX",
    "Estbdt",
    "OWNERSHIPTYPE",
    "Parval",
)

CSMAR_DAILY_FIELDS = (
    "Stkcd",
    "Trddt",
    "Opnprc",
    "Hiprc",
    "Loprc",
    "Clsprc",
    "Dnvaltrd",
    "Dsmvosd",
    "Dsmvtll",
    "Ahshrtrd_D",
)


@dataclass(frozen=True, slots=True)
class CSMARWorkbookEntry:
    archive_path: Path
    member_name: str
    archive_number: int
    member_number: int
    crc32: int
    file_size: int
    compressed_size: int

    @property
    def entry_id(self) -> str:
        return f"{self.archive_number:03d}_{self.member_number:03d}"

    @property
    def display_name(self) -> str:
        return f"{self.archive_path.name}:{self.member_name}"


@dataclass(frozen=True, slots=True)
class CSMARSourceLayout:
    root: Path
    master_path: Path
    daily_entries: tuple[CSMARWorkbookEntry, ...]
    fingerprint: str


class CSMARLocalReader:
    """Strict, offline reader for the exact directory selected by the user."""

    def __init__(self, source_root: str | Path) -> None:
        self.layout = discover_csmar_layout(source_root)

    def read_security_master(
        self,
        *,
        as_of: date,
        retrieved_at: datetime,
    ) -> pd.DataFrame:
        rows = list(_iter_local_xlsx_rows(self.layout.master_path, CSMAR_MASTER_FIELDS))
        if not rows:
            raise DataUnavailableError("CSMAR TRD_Co.xlsx 没有公司记录")
        raw = pd.DataFrame(rows, columns=list(CSMAR_MASTER_FIELDS))
        return normalize_csmar_security_master(
            raw,
            as_of=as_of,
            retrieved_at=retrieved_at,
        )

    def iter_daily_batches(
        self,
        entry: CSMARWorkbookEntry,
        *,
        batch_size: int,
    ) -> Iterator[pd.DataFrame]:
        if batch_size <= 0:
            raise ValueError("batch_size必须大于0")
        rows: list[tuple[object, ...]] = []
        with _open_nested_workbook(entry) as workbook:
            for row in _iter_workbook_rows(workbook, CSMAR_DAILY_FIELDS):
                rows.append(row)
                if len(rows) >= batch_size:
                    yield pd.DataFrame(rows, columns=list(CSMAR_DAILY_FIELDS))
                    rows = []
        if rows:
            yield pd.DataFrame(rows, columns=list(CSMAR_DAILY_FIELDS))


class CSMARParquetMarketData:
    """Query the completed local import through its DuckDB catalog."""

    def __init__(self, dataset_root: str | Path) -> None:
        self.root = Path(dataset_root).expanduser().resolve()
        self.database_path = self.root / "csmar.duckdb"

    def fetch_daily(
        self,
        symbol: str,
        start: date,
        end: date,
        *,
        adjust: str = "none",
    ) -> pd.DataFrame:
        if start > end:
            raise ValueError("start不能晚于end")
        if adjust.strip().lower() != "none":
            raise DataUnavailableError("CSMAR本地导出是未复权价格，不支持静默切换复权口径")
        normalized = normalize_symbol(symbol)
        connection = self._connect(read_only=True)
        try:
            result = connection.execute(
                "SELECT trade_date, open, high, low, close, prev_close, "
                "volume_shares, amount_cny, turnover_pct, source, retrieved_at "
                "FROM daily_bars WHERE symbol = ? AND trade_date BETWEEN ? AND ? "
                "ORDER BY trade_date",
                [normalized, start, end],
            ).fetchdf()
        finally:
            connection.close()
        if result.empty:
            raise DataUnavailableError(
                f"CSMAR本地库在{start.isoformat()}至{end.isoformat()}没有{normalized}日线"
            )
        result["trade_date"] = pd.to_datetime(result["trade_date"]).dt.normalize()
        result = result.loc[:, list(CANONICAL_DAILY_COLUMNS)]
        result.attrs.update(
            {
                "adjustment": "none",
                "data_quality": "licensed_local_export",
                "full_day_volume_available": False,
                "warning": "CSMAR_EXPORT_HAS_NO_FULL_DAY_VOLUME",
            }
        )
        return result

    def read_security_master(self, *, as_of: date | None = None) -> pd.DataFrame:
        connection = self._connect(read_only=True)
        try:
            if as_of is None:
                result = connection.execute(
                    "SELECT * FROM security_master ORDER BY symbol"
                ).fetchdf()
            else:
                result = connection.execute(
                    "SELECT * FROM security_master WHERE list_date <= ? ORDER BY symbol",
                    [as_of],
                ).fetchdf()
        finally:
            connection.close()
        if result.empty:
            raise DataUnavailableError("CSMAR本地证券主表为空")
        return result

    def read_full_market_for_date(self, trade_date: date) -> pd.DataFrame:
        connection = self._connect(read_only=True)
        try:
            result = connection.execute(
                "SELECT * FROM daily_bars WHERE trade_date = ? ORDER BY symbol",
                [trade_date],
            ).fetchdf()
        finally:
            connection.close()
        if result.empty:
            raise DataUnavailableError(f"CSMAR本地库没有{trade_date.isoformat()}全市场日线")
        result["trade_date"] = pd.to_datetime(result["trade_date"]).dt.normalize()
        return result

    def _connect(self, *, read_only: bool):
        if not self.database_path.is_file():
            raise DataUnavailableError(f"CSMAR DuckDB目录不存在：{self.database_path}")
        try:
            import duckdb
        except ImportError as exc:
            raise DataUnavailableError("缺少duckdb依赖，无法查询CSMAR本地库") from exc
        return duckdb.connect(str(self.database_path), read_only=read_only)


def discover_csmar_layout(source_root: str | Path) -> CSMARSourceLayout:
    root = Path(source_root).expanduser().resolve()
    if not root.is_dir():
        raise DataUnavailableError(f"CSMAR目录不存在：{root}")

    masters = sorted(root.rglob("TRD_Co.xlsx"))
    if len(masters) != 1:
        raise DataQualityError(f"CSMAR目录必须且只能包含一个TRD_Co.xlsx，实际找到{len(masters)}个")

    archives: list[tuple[int, Path]] = []
    for path in root.glob("日个股回报率文件*.zip"):
        match = re.fullmatch(r"日个股回报率文件(\d+)\.zip", path.name)
        if match:
            archives.append((int(match.group(1)), path.resolve()))
    archives.sort(key=lambda item: item[0])
    if not archives:
        raise DataUnavailableError("CSMAR目录没有日个股回报率文件*.zip")

    entries: list[CSMARWorkbookEntry] = []
    for archive_number, archive_path in archives:
        try:
            with zipfile.ZipFile(archive_path) as bundle:
                infos = [
                    info
                    for info in bundle.infolist()
                    if re.fullmatch(r"TRD_Dalyr\d*\.xlsx", Path(info.filename).name)
                ]
        except zipfile.BadZipFile as exc:
            raise DataQualityError(f"CSMAR压缩包损坏：{archive_path}") from exc
        if not infos:
            raise DataQualityError(f"{archive_path.name}没有TRD_Dalyr*.xlsx")
        infos.sort(key=lambda item: _daily_member_number(Path(item.filename).name))
        for info in infos:
            member_number = _daily_member_number(Path(info.filename).name)
            entries.append(
                CSMARWorkbookEntry(
                    archive_path=archive_path,
                    member_name=info.filename,
                    archive_number=archive_number,
                    member_number=member_number,
                    crc32=info.CRC,
                    file_size=info.file_size,
                    compressed_size=info.compress_size,
                )
            )

    fingerprint_payload = {
        "schema": 1,
        "master": {
            "relative_path": str(masters[0].relative_to(root)),
            "size": masters[0].stat().st_size,
            "sha256": _sha256_file(masters[0]),
        },
        "daily_entries": [
            {
                "archive": str(entry.archive_path.relative_to(root)),
                "member": entry.member_name,
                "crc32": entry.crc32,
                "file_size": entry.file_size,
                "compressed_size": entry.compressed_size,
            }
            for entry in entries
        ],
    }
    fingerprint = hashlib.sha256(
        json.dumps(fingerprint_payload, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return CSMARSourceLayout(
        root=root,
        master_path=masters[0].resolve(),
        daily_entries=tuple(entries),
        fingerprint=fingerprint,
    )


def normalize_csmar_security_master(
    raw: pd.DataFrame,
    *,
    as_of: date,
    retrieved_at: datetime,
) -> pd.DataFrame:
    missing = set(CSMAR_MASTER_FIELDS) - set(raw.columns)
    if missing:
        raise DataQualityError("CSMAR公司文件缺少字段：" + ", ".join(sorted(missing)))
    if retrieved_at.tzinfo is None or retrieved_at.utcoffset() is None:
        raise ValueError("retrieved_at必须包含时区")

    frame = raw.loc[:, list(CSMAR_MASTER_FIELDS)].copy()
    frame["symbol"] = frame["Stkcd"].map(_normalize_code)
    frame["exchange"] = frame["symbol"].map(infer_a_share_exchange)
    frame = frame.loc[frame["exchange"].notna()].copy()
    if frame.empty:
        raise DataUnavailableError("CSMAR公司文件没有可识别的A股证券")

    frame["list_date"] = pd.to_datetime(frame["Listdt"], errors="coerce").dt.normalize()
    if bool(frame["list_date"].isna().any()):
        examples = frame.loc[frame["list_date"].isna(), "symbol"].head(5).tolist()
        raise DataQualityError(f"CSMAR公司文件存在无效上市日期：{examples}")
    future = frame["list_date"] > pd.Timestamp(as_of)
    frame = frame.loc[~future].copy()

    frame["name"] = frame["Stknme"].fillna("").astype(str).str.strip()
    if bool(frame["name"].eq("").any()):
        raise DataQualityError("CSMAR公司文件存在空证券简称")
    if bool(frame["symbol"].duplicated().any()):
        duplicates = frame.loc[frame["symbol"].duplicated(keep=False), "symbol"].unique()
        raise DataQualityError(f"CSMAR公司文件存在重复证券代码：{duplicates[:5].tolist()}")

    normalized_name = frame["name"].str.upper().str.replace(" ", "", regex=False)
    industry = frame["IndnmeZX"].fillna(frame["Nnindnme"]).fillna("")
    established = frame["Estbdt"].astype(str).str.replace(r"-00$", "-01", regex=True)
    established = established.str.replace(r"-00-", "-01-", regex=True)
    timestamp = retrieved_at.astimezone(UTC).isoformat().replace("+00:00", "Z")

    output = pd.DataFrame(
        {
            "symbol": frame["symbol"],
            "name": frame["name"],
            "exchange": frame["exchange"],
            "board": frame["symbol"].map(infer_a_share_board),
            "list_date": frame["list_date"],
            "delist_date": pd.NaT,
            "industry": industry.astype(str).str.strip(),
            "is_st": normalized_name.str.startswith(("ST", "*ST", "PT")),
            "is_delisting": normalized_name.str.contains("退", regex=False),
            "is_suspended": False,
            "source": "csmar",
            "retrieved_at": timestamp,
            "industry_csrc_2012": frame["Nnindnme"].fillna("").astype(str).str.strip(),
            "industry_listing_association": frame["IndnmeZX"].fillna("").astype(str).str.strip(),
            "established_date": pd.to_datetime(established, errors="coerce").dt.normalize(),
            "ownership_type": frame["OWNERSHIPTYPE"].fillna("").astype(str).str.strip(),
            "par_value_cny": pd.to_numeric(frame["Parval"], errors="coerce"),
        }
    )
    return output.sort_values("symbol").reset_index(drop=True)


def infer_a_share_exchange(symbol: str) -> str | None:
    """Return the exchange for common A-share company-code ranges.

    Company exports contain B shares as well.  ``2xxxxx`` (Shenzhen B) and
    ``9xxxxx`` other than the Beijing ``920xxx`` range are deliberately
    excluded instead of being mislabelled as A shares.
    """

    if not (len(symbol) == 6 and symbol.isdigit()):
        return None
    if symbol.startswith("6"):
        return "SH"
    if symbol.startswith(("0", "3")):
        return "SZ"
    if symbol.startswith(("4", "8", "92")):
        return "BJ"
    return None


def infer_a_share_board(symbol: str) -> str:
    exchange = infer_a_share_exchange(symbol)
    if exchange == "BJ":
        return "北京证券交易所"
    if symbol.startswith(("688", "689")):
        return "科创板"
    if symbol.startswith(("300", "301")):
        return "创业板"
    return "主板"


def _normalize_code(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text.zfill(6) if text.isdigit() and len(text) <= 6 else text


def _daily_member_number(name: str) -> int:
    match = re.fullmatch(r"TRD_Dalyr(\d*)\.xlsx", name)
    if not match:
        raise ValueError(f"不是CSMAR日线工作簿：{name}")
    suffix = match.group(1)
    return int(suffix) if suffix else 0


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while block := handle.read(1024 * 1024):
            digest.update(block)
    return digest.hexdigest()


def _iter_local_xlsx_rows(
    path: Path,
    required_fields: tuple[str, ...],
) -> Iterator[tuple[object, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        yield from _iter_workbook_rows(workbook, required_fields)
    finally:
        workbook.close()


@contextmanager
def _open_nested_workbook(entry: CSMARWorkbookEntry):
    """Stage one nested XLSX in a bounded spool; the source ZIP stays read-only."""

    with zipfile.ZipFile(entry.archive_path) as bundle:
        try:
            source: BinaryIO = bundle.open(entry.member_name)
        except KeyError as exc:
            raise DataQualityError(f"压缩包成员消失：{entry.display_name}") from exc
        with source, tempfile.SpooledTemporaryFile(max_size=64 * 1024 * 1024) as staged:
            shutil.copyfileobj(source, staged, length=1024 * 1024)
            staged.seek(0)
            workbook = load_workbook(staged, read_only=True, data_only=True)
            try:
                yield workbook
            finally:
                workbook.close()


def _iter_workbook_rows(workbook, required_fields: tuple[str, ...]):
    worksheet = workbook.active
    # Required for CSMAR exports whose XML dimension incorrectly says A1:A... .
    worksheet.reset_dimensions()
    iterator = worksheet.iter_rows(values_only=True)
    try:
        header = tuple(str(value).strip() if value is not None else "" for value in next(iterator))
    except StopIteration as exc:
        raise DataUnavailableError("CSMAR工作簿为空") from exc
    missing = set(required_fields) - set(header)
    if missing:
        raise DataQualityError("CSMAR工作簿缺少字段：" + ", ".join(sorted(missing)))
    positions = tuple(header.index(field) for field in required_fields)

    # CSMAR exports place Chinese labels and units in rows 2 and 3.
    try:
        next(iterator)
        next(iterator)
    except StopIteration as exc:
        raise DataUnavailableError("CSMAR工作簿只有表头，没有数据") from exc
    for row in iterator:
        if not row or all(value is None for value in row):
            continue
        yield tuple(row[position] if position < len(row) else None for position in positions)
